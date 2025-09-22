import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO


class ExcelExporter:
    """Classe utilitaire pour l'export Excel"""
    
    def __init__(self, title="Export"):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = title
        self.current_row = 1
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def add_title(self, title, subtitle=None):
        """Ajoute un titre principal"""
        # Titre principal
        self.worksheet.merge_cells(f'A{self.current_row}:Z{self.current_row}')
        title_cell = self.worksheet.cell(row=self.current_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=16, color="366092")
        title_cell.alignment = Alignment(horizontal="center")
        self.current_row += 1
        
        # Sous-titre
        if subtitle:
            self.worksheet.merge_cells(f'A{self.current_row}:Z{self.current_row}')
            subtitle_cell = self.worksheet.cell(row=self.current_row, column=1, value=subtitle)
            subtitle_cell.font = Font(size=12, color="666666")
            subtitle_cell.alignment = Alignment(horizontal="center")
            self.current_row += 1
        
        # Date d'export
        date_text = f"Exporté le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        self.worksheet.merge_cells(f'A{self.current_row}:Z{self.current_row}')
        date_cell = self.worksheet.cell(row=self.current_row, column=1, value=date_text)
        date_cell.font = Font(size=10, color="999999")
        date_cell.alignment = Alignment(horizontal="center")
        self.current_row += 2
    
    def add_headers(self, headers):
        """Ajoute les en-têtes de colonnes"""
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        self.current_row += 1
        return self.current_row - 1  # Retourne la ligne des headers
    
    def add_data_row(self, data):
        """Ajoute une ligne de données"""
        for col, value in enumerate(data, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=value)
            cell.border = self.border
            
            # Formatage conditionnel selon le type de données
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, str) and value.startswith('http'):
                cell.style = 'Hyperlink'
        
        self.current_row += 1
    
    def add_summary_section(self, title, data):
        """Ajoute une section de résumé"""
        self.current_row += 1
        
        # Titre de la section
        title_cell = self.worksheet.cell(row=self.current_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color="366092")
        self.current_row += 1
        
        # Données du résumé
        for key, value in data.items():
            self.worksheet.cell(row=self.current_row, column=1, value=key)
            self.worksheet.cell(row=self.current_row, column=2, value=value)
            self.current_row += 1
    
    def auto_adjust_columns(self):
        """Ajuste automatiquement la largGNF des colonnes"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_chart(self, chart_type, data_range, title="Graphique"):
        """Ajoute un graphique (basique)"""
        from openpyxl.chart import BarChart, Reference
        
        if chart_type == 'bar':
            chart = BarChart()
            chart.title = title
            chart.x_axis.title = 'Catégories'
            chart.y_axis.title = 'Valeurs'
            
            # Ajouter les données
            data = Reference(self.worksheet, min_col=2, min_row=data_range[0], 
                           max_row=data_range[1], max_col=2)
            categories = Reference(self.worksheet, min_col=1, min_row=data_range[0], 
                                 max_row=data_range[1])
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Positionner le graphique
            self.worksheet.add_chart(chart, f"E{self.current_row}")
    
    def get_response(self, filename):
        """Retourne la réponse HTTP avec le fichier Excel"""
        # Ajuster les colonnes
        self.auto_adjust_columns()
        
        # Créer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le workbook
        self.workbook.save(response)
        return response


def export_applications_to_excel():
    """Exporte les candidatures vers Excel"""
    from apps.applications.models import Application
    
    exporter = ExcelExporter("Candidatures")
    
    # Titre
    exporter.add_title(
        "Rapport des Candidatures",
        "Export complet de toutes les candidatures"
    )
    
    # En-têtes
    headers = [
        'ID', 'Candidat', 'Email', 'Téléphone', 'Offre', 'Entreprise',
        'Statut', 'Priorité', 'Date candidature', 'Salaire souhaité',
        'Disponibilité', 'Mobilité', 'Examiné par', 'Date examen'
    ]
    exporter.add_headers(headers)
    
    # Données
    applications = Application.objects.select_related(
        'candidate__user', 'job', 'reviewed_by'
    ).all()
    
    for app in applications:
        data = [
            app.id,
            app.candidate.user.full_name,
            app.candidate.user.email,
            app.candidate.mobile_phone or '',
            app.job.title,
            app.job.company,
            app.get_status_display(),
            app.get_priority_display(),
            app.applied_at.strftime('%d/%m/%Y %H:%M'),
            f"{app.expected_salary} €" if app.expected_salary else '',
            app.availability_date.strftime('%d/%m/%Y') if app.availability_date else '',
            'Oui' if app.willing_to_relocate else 'Non',
            app.reviewed_by.full_name if app.reviewed_by else '',
            app.reviewed_at.strftime('%d/%m/%Y %H:%M') if app.reviewed_at else ''
        ]
        exporter.add_data_row(data)
    
    # Statistiques
    stats = {
        'Total candidatures': applications.count(),
        'En attente': applications.filter(status='pending').count(),
        'En cours d\'examen': applications.filter(status='reviewing').count(),
        'Présélectionnés': applications.filter(status='shortlisted').count(),
        'Acceptés': applications.filter(status='accepted').count(),
        'Rejetés': applications.filter(status='rejected').count(),
    }
    exporter.add_summary_section("Statistiques", stats)
    
    filename = f'candidatures_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return exporter.get_response(filename)


def export_candidates_to_excel():
    """Exporte les candidats vers Excel"""
    from apps.accounts.models import CandidateProfile
    
    exporter = ExcelExporter("Candidats")
    
    # Titre
    exporter.add_title(
        "Base de Données Candidats",
        "Export complet de tous les profils candidats"
    )
    
    # En-têtes
    headers = [
        'ID', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays',
        'Poste actuel', 'Entreprise', 'Expérience (années)', 'Salaire souhaité',
        'Type poste', 'Mobilité', 'Profil (%)', 'Inscription', 'Actif'
    ]
    exporter.add_headers(headers)
    
    # Données
    candidates = CandidateProfile.objects.select_related('user').all()
    
    for candidate in candidates:
        data = [
            candidate.id,
            candidate.user.first_name,
            candidate.user.last_name,
            candidate.user.email,
            candidate.mobile_phone or '',
            candidate.city or '',
            candidate.country or '',
            candidate.current_position or '',
            candidate.current_company or '',
            candidate.years_of_experience,
            f"{candidate.expected_salary} €" if candidate.expected_salary else '',
            candidate.get_preferred_work_type_display() if candidate.preferred_work_type else '',
            'Oui' if candidate.willing_to_relocate else 'Non',
            f"{candidate.profile_completion}%",
            candidate.created_at.strftime('%d/%m/%Y'),
            'Oui' if candidate.is_active else 'Non'
        ]
        exporter.add_data_row(data)
    
    # Statistiques
    stats = {
        'Total candidats': candidates.count(),
        'Profils actifs': candidates.filter(is_active=True).count(),
        'Avec CV': candidates.exclude(cv_file='').count(),
        'Profil > 80%': candidates.filter(profile_completion__gte=80).count(),
        'Mobiles': candidates.filter(willing_to_relocate=True).count(),
    }
    exporter.add_summary_section("Statistiques", stats)
    
    filename = f'candidats_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return exporter.get_response(filename)


def export_jobs_to_excel():
    """Exporte les offres d'emploi vers Excel"""
    from apps.jobs.models import Job
    
    exporter = ExcelExporter("Offres d'emploi")
    
    # Titre
    exporter.add_title(
        "Catalogue des Offres d'Emploi",
        "Export complet de toutes les offres d'emploi"
    )
    
    # En-têtes
    headers = [
        'ID', 'Titre', 'Entreprise', 'Catégorie', 'Type', 'Niveau',
        'Localisation', 'Télétravail', 'Salaire min', 'Salaire max',
        'Statut', 'Vedette', 'Urgent', 'Candidatures', 'Vues',
        'Création', 'Limite', 'Créé par'
    ]
    exporter.add_headers(headers)
    
    # Données
    jobs = Job.objects.select_related('category', 'created_by').all()
    
    for job in jobs:
        data = [
            job.id,
            job.title,
            job.company,
            job.category.name,
            job.get_job_type_display(),
            job.get_experience_level_display(),
            job.location,
            'Oui' if job.remote_work else 'Non',
            f"{job.salary_min} €" if job.salary_min else '',
            f"{job.salary_max} €" if job.salary_max else '',
            job.get_status_display(),
            'Oui' if job.featured else 'Non',
            'Oui' if job.urgent else 'Non',
            job.applications_count,
            job.views_count,
            job.created_at.strftime('%d/%m/%Y'),
            job.application_deadline.strftime('%d/%m/%Y') if job.application_deadline else '',
            job.created_by.full_name
        ]
        exporter.add_data_row(data)
    
    # Statistiques
    stats = {
        'Total offres': jobs.count(),
        'Publiées': jobs.filter(status='published').count(),
        'En vedette': jobs.filter(featured=True).count(),
        'Urgentes': jobs.filter(urgent=True).count(),
        'Télétravail': jobs.filter(remote_work=True).count(),
    }
    exporter.add_summary_section("Statistiques", stats)
    
    filename = f'offres_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return exporter.get_response(filename)

