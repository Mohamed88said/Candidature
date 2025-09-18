from django.http import HttpResponse
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from apps.accounts.models import User, CandidateProfile
from apps.jobs.models import Job
from apps.applications.models import Application, Interview


def get_dashboard_stats():
    """Récupère les statistiques pour le dashboard"""
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # Statistiques générales
    stats = {
        'total_candidates': CandidateProfile.objects.count(),
        'total_jobs': Job.objects.filter(status='published').count(),
        'total_applications': Application.objects.count(),
        'total_interviews': Interview.objects.count(),
        
        # Nouvelles cette semaine
        'new_candidates_week': CandidateProfile.objects.filter(created_at__date__gte=week_ago).count(),
        'new_applications_week': Application.objects.filter(applied_at__date__gte=week_ago).count(),
        'new_jobs_week': Job.objects.filter(created_at__date__gte=week_ago).count(),
        
        # Applications par statut
        'pending_applications': Application.objects.filter(status='pending').count(),
        'reviewing_applications': Application.objects.filter(status='reviewing').count(),
        'shortlisted_applications': Application.objects.filter(status='shortlisted').count(),
        'interview_scheduled': Application.objects.filter(status='interview_scheduled').count(),
        'offers_made': Application.objects.filter(status='offer_made').count(),
        'accepted_applications': Application.objects.filter(status='accepted').count(),
        'rejected_applications': Application.objects.filter(status='rejected').count(),
        
        # Entretiens
        'upcoming_interviews': Interview.objects.filter(
            scheduled_date__gte=timezone.now(),
            status='scheduled'
        ).count(),
        'interviews_today': Interview.objects.filter(
            scheduled_date__date=today,
            status='scheduled'
        ).count(),
        
        # Taux de conversion
        'conversion_rate': 0,
        'avg_time_to_hire': 0,
    }
    
    # Calcul du taux de conversion
    total_apps = stats['total_applications']
    if total_apps > 0:
        stats['conversion_rate'] = round((stats['accepted_applications'] / total_apps) * 100, 1)
    
    # Temps moyen d'embauche (en jours)
    hired_apps = Application.objects.filter(status='accepted')
    if hired_apps.exists():
        total_days = sum([
            (app.updated_at.date() - app.applied_at.date()).days 
            for app in hired_apps
        ])
        stats['avg_time_to_hire'] = round(total_days / hired_apps.count(), 1)
    
    return stats


def generate_excel_report(report_type):
    """Génère un rapport Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if report_type == 'applications':
        ws.title = "Candidatures"
        
        # En-têtes
        headers = [
            'ID', 'Candidat', 'Email', 'Téléphone', 'Offre', 'Entreprise',
            'Statut', 'Priorité', 'Date de candidature', 'Salaire souhaité',
            'Date de disponibilité', 'Mobilité', 'Examiné par', 'Date d\'examen'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        applications = Application.objects.select_related(
            'candidate__user', 'job', 'reviewed_by'
        ).all()
        
        for row, app in enumerate(applications, 2):
            ws.cell(row=row, column=1, value=app.id)
            ws.cell(row=row, column=2, value=app.candidate.user.full_name)
            ws.cell(row=row, column=3, value=app.candidate.user.email)
            ws.cell(row=row, column=4, value=app.candidate.mobile_phone or '')
            ws.cell(row=row, column=5, value=app.job.title)
            ws.cell(row=row, column=6, value=app.job.company)
            ws.cell(row=row, column=7, value=app.get_status_display())
            ws.cell(row=row, column=8, value=app.get_priority_display())
            ws.cell(row=row, column=9, value=app.applied_at.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=10, value=str(app.expected_salary) if app.expected_salary else '')
            ws.cell(row=row, column=11, value=app.availability_date.strftime('%d/%m/%Y') if app.availability_date else '')
            ws.cell(row=row, column=12, value='Oui' if app.willing_to_relocate else 'Non')
            ws.cell(row=row, column=13, value=app.reviewed_by.full_name if app.reviewed_by else '')
            ws.cell(row=row, column=14, value=app.reviewed_at.strftime('%d/%m/%Y %H:%M') if app.reviewed_at else '')
    
    elif report_type == 'candidates':
        ws.title = "Candidats"
        
        # En-têtes
        headers = [
            'ID', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Ville', 'Pays',
            'Poste actuel', 'Entreprise actuelle', 'Années d\'expérience',
            'Salaire souhaité', 'Type de poste', 'Mobilité', 'Completion profil (%)',
            'Date d\'inscription', 'Actif'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        candidates = CandidateProfile.objects.select_related('user').all()
        
        for row, candidate in enumerate(candidates, 2):
            ws.cell(row=row, column=1, value=candidate.id)
            ws.cell(row=row, column=2, value=candidate.user.first_name)
            ws.cell(row=row, column=3, value=candidate.user.last_name)
            ws.cell(row=row, column=4, value=candidate.user.email)
            ws.cell(row=row, column=5, value=candidate.mobile_phone or '')
            ws.cell(row=row, column=6, value=candidate.city or '')
            ws.cell(row=row, column=7, value=candidate.country or '')
            ws.cell(row=row, column=8, value=candidate.current_position or '')
            ws.cell(row=row, column=9, value=candidate.current_company or '')
            ws.cell(row=row, column=10, value=candidate.years_of_experience)
            ws.cell(row=row, column=11, value=str(candidate.expected_salary) if candidate.expected_salary else '')
            ws.cell(row=row, column=12, value=candidate.get_preferred_work_type_display() if candidate.preferred_work_type else '')
            ws.cell(row=row, column=13, value='Oui' if candidate.willing_to_relocate else 'Non')
            ws.cell(row=row, column=14, value=candidate.profile_completion)
            ws.cell(row=row, column=15, value=candidate.created_at.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=16, value='Oui' if candidate.is_active else 'Non')
    
    elif report_type == 'jobs':
        ws.title = "Offres d'emploi"
        
        # En-têtes
        headers = [
            'ID', 'Titre', 'Entreprise', 'Catégorie', 'Type', 'Niveau d\'expérience',
            'Localisation', 'Télétravail', 'Salaire min', 'Salaire max', 'Devise',
            'Statut', 'En vedette', 'Urgent', 'Nb candidatures', 'Nb vues',
            'Date de création', 'Date limite', 'Créé par'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        jobs = Job.objects.select_related('category', 'created_by').all()
        
        for row, job in enumerate(jobs, 2):
            ws.cell(row=row, column=1, value=job.id)
            ws.cell(row=row, column=2, value=job.title)
            ws.cell(row=row, column=3, value=job.company)
            ws.cell(row=row, column=4, value=job.category.name)
            ws.cell(row=row, column=5, value=job.get_job_type_display())
            ws.cell(row=row, column=6, value=job.get_experience_level_display())
            ws.cell(row=row, column=7, value=job.location)
            ws.cell(row=row, column=8, value='Oui' if job.remote_work else 'Non')
            ws.cell(row=row, column=9, value=str(job.salary_min) if job.salary_min else '')
            ws.cell(row=row, column=10, value=str(job.salary_max) if job.salary_max else '')
            ws.cell(row=row, column=11, value=job.salary_currency)
            ws.cell(row=row, column=12, value=job.get_status_display())
            ws.cell(row=row, column=13, value='Oui' if job.featured else 'Non')
            ws.cell(row=row, column=14, value='Oui' if job.urgent else 'Non')
            ws.cell(row=row, column=15, value=job.applications_count)
            ws.cell(row=row, column=16, value=job.views_count)
            ws.cell(row=row, column=17, value=job.created_at.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=18, value=job.application_deadline.strftime('%d/%m/%Y %H:%M') if job.application_deadline else '')
            ws.cell(row=row, column=19, value=job.created_by.full_name)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    wb.save(response)
    
    return response


def calculate_candidate_score(candidate):
    """Calcule un score pour un candidat basé sur son profil"""
    score = 0
    
    # Completion du profil (0-30 points)
    score += (candidate.profile_completion / 100) * 30
    
    # Expérience (0-25 points)
    exp_score = min(candidate.years_of_experience * 2.5, 25)
    score += exp_score
    
    # Formation (0-20 points)
    education_count = candidate.educations.count()
    education_score = min(education_count * 5, 20)
    score += education_score
    
    # Compétences (0-15 points)
    skills_count = candidate.skills.count()
    skills_score = min(skills_count * 1.5, 15)
    score += skills_score
    
    # Documents (0-10 points)
    if candidate.cv_file:
        score += 5
    if candidate.cover_letter:
        score += 5
    
    return min(round(score, 1), 100)


def get_recruitment_analytics():
    """Analyse avancée des données de recrutement"""
    analytics = {}
    
    # Analyse des sources de candidatures (si implémenté)
    # analytics['sources'] = Application.objects.values('source').annotate(count=Count('id'))
    
    # Analyse des délais de recrutement
    hired_apps = Application.objects.filter(status='accepted')
    if hired_apps.exists():
        hiring_times = []
        for app in hired_apps:
            days = (app.updated_at.date() - app.applied_at.date()).days
            hiring_times.append(days)
        
        analytics['hiring_time'] = {
            'avg': sum(hiring_times) / len(hiring_times),
            'min': min(hiring_times),
            'max': max(hiring_times),
            'median': sorted(hiring_times)[len(hiring_times)//2]
        }
    
    # Analyse des taux de réussite par catégorie
    categories_success = []
    for category in Job.objects.values('category__name').distinct():
        cat_name = category['category__name']
        total_apps = Application.objects.filter(job__category__name=cat_name).count()
        hired_apps = Application.objects.filter(
            job__category__name=cat_name, 
            status='accepted'
        ).count()
        
        if total_apps > 0:
            success_rate = (hired_apps / total_apps) * 100
            categories_success.append({
                'category': cat_name,
                'success_rate': round(success_rate, 1),
                'total_applications': total_apps
            })
    
    analytics['categories_success'] = sorted(
        categories_success, 
        key=lambda x: x['success_rate'], 
        reverse=True
    )
    
    return analytics